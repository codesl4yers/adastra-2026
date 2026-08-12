"""Lee el índice maestro que entrega ADL con el corpus.

Es la fuente de verdad de la identidad de cada documento: ``DOC_ID``, fenómeno y
observatorio vienen de aquí. Este módulo **solo lee**; la escritura es exclusiva
de :mod:`orquestador`.

Por qué manda el índice y no la deducción por nombre o carpeta:
``docs/decisiones/orquestacion-y-determinismo.md``.

Uso::

    from indice import cargar_indice
    entradas = cargar_indice(Path("Indice_Datos_Codefest.xlsx"))
    entradas["F1_IA_y_Capacidades_Estrategicas/.../informe.pdf"].doc_id
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

HOJA = "Inventario de Archivos"

_COL_FENOMENO = "Fenómeno"
_COL_OBSERVATORIO = "Observatorio"
_COL_CODIGO = "Código Observatorio"
_COL_DOC_ID = "DOC_ID"
_COL_NOMBRE = "Nombre estandarizado"
_COL_CARPETA = "Carpeta"
_COL_TIPO = "Tipo"

_COLUMNAS: tuple[str, ...] = (
    _COL_FENOMENO,
    _COL_OBSERVATORIO,
    _COL_CODIGO,
    _COL_DOC_ID,
    _COL_NOMBRE,
    _COL_CARPETA,
    _COL_TIPO,
)

# La columna "Carpeta" puede venir vacía si el archivo está en la raíz.
_COLUMNAS_OPCIONALES: frozenset[str] = frozenset({_COL_CARPETA})

_FENOMENOS: dict[str, int] = {"F1": 1, "F2": 2, "F3": 3}

# El DOC_ID se usa crudo como nombre de archivo. No se exige la forma completa
# F<n>-<CODIGO>-<nnn>: el número de dígitos no es parte del contrato.
_CARACTERES_PROHIBIDOS_EN_DOC_ID = frozenset('/\\:*?"<>|')


@dataclass(frozen=True)
class EntradaIndice:
    """Una fila del inventario de ADL, ya normalizada."""

    doc_id: str               # "F1-AIINDEX-001"
    fuente: str               # "Nombre estandarizado", el nombre exacto del archivo
    ruta_relativa: str        # "Carpeta/Nombre" POSIX; es la clave del mapa
    fenomeno: int             # 1, 2 o 3
    observatorio: str         # "CSET_Georgetown"
    codigo_observatorio: str  # "CSET"
    tipo_declarado: str       # lo que declara ADL: PDF, JSON, Otro...


def cargar_indice(ruta_xlsx: Path) -> dict[str, EntradaIndice]:
    """Devuelve un mapa ``ruta_relativa -> EntradaIndice``, en el orden del archivo.

    La clave es la ruta y no el nombre porque el nombre no es único: 59 se
    repiten en 186 filas. Lanza ``ValueError`` si el índice es inconsistente
    —``DOC_ID`` o rutas repetidas, columnas ausentes, fenómeno fuera de rango—.
    """
    ruta_xlsx = Path(ruta_xlsx)
    if not ruta_xlsx.is_file():
        raise ValueError(f"el índice no existe: {ruta_xlsx}")

    libro = openpyxl.load_workbook(ruta_xlsx, read_only=True, data_only=True)
    try:
        if HOJA not in libro.sheetnames:
            raise ValueError(
                f"el índice no tiene la hoja {HOJA!r}; tiene {libro.sheetnames}"
            )
        return _leer_hoja(libro[HOJA])
    finally:
        libro.close()


def _leer_hoja(hoja) -> dict[str, EntradaIndice]:
    """Recorre la hoja fila a fila.

    Sin ``ws.max_row``: en ``read_only`` cuenta filas con formato y sin datos.
    """
    filas = hoja.iter_rows(values_only=True)
    try:
        cabecera = next(filas)
    except StopIteration:
        raise ValueError(f"la hoja {HOJA!r} está vacía") from None

    posiciones = _posiciones_de_columnas(cabecera)

    entradas: dict[str, EntradaIndice] = {}
    doc_ids: dict[str, str] = {}

    # La cabecera es la fila 1, así que los datos empiezan en la 2.
    for numero, fila in enumerate(filas, start=2):
        if all(celda is None for celda in fila):
            continue

        entrada = _entrada_de_fila(fila, posiciones, numero)

        anterior = entradas.get(entrada.ruta_relativa)
        if anterior is not None:
            raise ValueError(
                f"fila {numero}: ruta duplicada {entrada.ruta_relativa!r} "
                f"(ya la usa {anterior.doc_id}). La ruta es la clave de join "
                f"y debe ser única."
            )

        ruta_previa = doc_ids.get(entrada.doc_id)
        if ruta_previa is not None:
            raise ValueError(
                f"fila {numero}: DOC_ID duplicado {entrada.doc_id!r} "
                f"({ruta_previa} y {entrada.ruta_relativa}). "
                f"Un índice con identidades repetidas invalida la trazabilidad."
            )

        entradas[entrada.ruta_relativa] = entrada
        doc_ids[entrada.doc_id] = entrada.ruta_relativa

    return entradas


def _posiciones_de_columnas(cabecera: tuple) -> dict[str, int]:
    """Mapea nombre de columna a su posición, para no depender del orden."""
    posiciones: dict[str, int] = {}
    for posicion, celda in enumerate(cabecera):
        if celda is None:
            continue
        posiciones[str(celda).strip()] = posicion

    faltan = [columna for columna in _COLUMNAS if columna not in posiciones]
    if faltan:
        raise ValueError(
            f"al índice le faltan columnas: {faltan}. "
            f"Esperadas: {list(_COLUMNAS)}. Encontradas: {sorted(posiciones)}"
        )
    return posiciones


def _celda(fila: tuple, posiciones: dict[str, int], columna: str, numero: int) -> str:
    """Lee una celda como texto, exigiendo que las obligatorias no estén vacías."""
    posicion = posiciones[columna]
    valor = fila[posicion] if posicion < len(fila) else None
    texto = "" if valor is None else str(valor).strip()

    if not texto and columna not in _COLUMNAS_OPCIONALES:
        raise ValueError(f"fila {numero}: la columna {columna!r} está vacía")
    return texto


def _entrada_de_fila(
    fila: tuple, posiciones: dict[str, int], numero: int
) -> EntradaIndice:
    bruto = _celda(fila, posiciones, _COL_FENOMENO, numero).upper()
    if bruto not in _FENOMENOS:
        raise ValueError(
            f"fila {numero}: fenómeno {bruto!r} no es F1, F2 ni F3"
        )

    doc_id = _celda(fila, posiciones, _COL_DOC_ID, numero)
    _validar_doc_id(doc_id, numero)

    nombre = _celda(fila, posiciones, _COL_NOMBRE, numero)
    # ADL genera el índice en Windows; el pipeline puede correr en Linux.
    carpeta = _celda(fila, posiciones, _COL_CARPETA, numero).replace("\\", "/").strip("/")

    return EntradaIndice(
        doc_id=doc_id,
        fuente=nombre,
        ruta_relativa=f"{carpeta}/{nombre}" if carpeta else nombre,
        fenomeno=_FENOMENOS[bruto],
        observatorio=_celda(fila, posiciones, _COL_OBSERVATORIO, numero),
        codigo_observatorio=_celda(fila, posiciones, _COL_CODIGO, numero),
        tipo_declarado=_celda(fila, posiciones, _COL_TIPO, numero),
    )


def _validar_doc_id(doc_id: str, numero: int) -> None:
    """Rechaza un ``DOC_ID`` que no sirva como nombre de archivo.

    Se comprueba aquí, antes de escribir nada: una barra por typo reventaría la
    escritura a mitad de una corrida de 1826 archivos.
    """
    prohibidos = sorted(_CARACTERES_PROHIBIDOS_EN_DOC_ID.intersection(doc_id))
    if prohibidos:
        raise ValueError(
            f"fila {numero}: DOC_ID {doc_id!r} no sirve como nombre de "
            f"archivo: contiene {prohibidos}"
        )

"""Extractor de CSV y XLSX: los datasets del corpus.

Un solo módulo para los dos formatos porque comparten el modelo de datos —filas
y columnas— y toda la lógica de serialización. Solo cambia el lector: el ``csv``
de la estándar para uno, ``openpyxl`` en modo ``read_only`` para el otro. El
``formato`` del ``Documento`` sí los distingue, por la extensión.

Cada fila es un bloque ``atomico``: un registro es una unidad de significado
completa. Partirlo rompe la correspondencia ``columna: valor`` y fusionar dos
produce un vector que no representa fielmente a ninguno de los dos. El nombre
de la columna se repite en cada fila —redundante para un humano— porque es lo
que permite que la fila se recupere sola, sin la cabecera como contexto.

La trampa del CSV
-----------------
Un CSV no declara ni su codificación ni su delimitador. Un archivo exportado
desde Excel en español viene en cp1252 y separado por ``;``: leerlo como UTF-8
con ``,`` produce **una única columna con todo el contenido dentro y ni un solo
error**. Aquí el delimitador se detecta con ``csv.Sniffer`` y la codificación
probando en un orden fijo (BOM, UTF-8, cp1252, latin-1). El orden es fijo a
propósito: una autodetección probabilística puede decodificar distinto en dos
corridas y romper la reproducibilidad.

La trampa del XLSX
------------------
Las celdas con fórmula devuelven la fórmula y no el valor si no se abre con
``data_only=True``, y las fechas llegan como enteros seriales de Excel. Además,
``openpyxl`` sin ``read_only=True`` carga la hoja entera en memoria.
"""

from __future__ import annotations

import csv
import datetime
import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from contrato import Bloque, Documento
from extractores.comun import Jerarquia, construir_documento, documento_fallido, serializar_registro
from limpieza import normalizar_texto

FORMATOS_SOPORTADOS = {".csv": "csv", ".xlsx": "xlsx"}

# Codificaciones a probar, en este orden. utf-8-sig va primero porque acierta
# también con los archivos sin BOM; latin-1 va última porque nunca falla y
# taparía a las demás.
CODIFICACIONES = ("utf-8-sig", "cp1252", "latin-1")

DELIMITADORES = ",;\t|"

# Tope de filas por archivo. Un dataset de decenas de miles de filas produce
# decenas de miles de vectores casi idénticos que ahogan el índice sin aportar
# recuperabilidad. Se trunca y se deja constancia; en silencio sería peor.
MAXIMO_FILAS = 5000

# Muestra que se le pasa al Sniffer. Con más no acierta más y con archivos de
# 1 MB leerlo entero para adivinar una coma es tiempo tirado.
MUESTRA_SNIFFER = 8192


@dataclass(frozen=True)
class EsquemaTabular:
    """Qué columnas de un export conocido merecen entrar al vector.

    Se identifica por la **cabecera**, no por el nombre del archivo: el nombre
    cambia entre descargas y la cabecera es el export en sí. La firma es un
    subconjunto requerido para tolerar la columna de índice sin nombre que
    algunos exports traen y otros no.
    """

    nombre: str
    firma: frozenset[str]
    indexables: tuple[str, ...]
    """En el orden en que se emiten. El resto de columnas va a metadata."""


# Los exports bibliográficos son el caso extremo del corpus: 264.264 filas
# entre siete archivos, y en cada una más de la mitad del texto son
# identificadores —PMID, DOI, PMCID, NIHMS ID— que no aportan al embedding y
# diluyen el título, que es lo único que se puede recuperar por semejanza.
# Se declaran a mano y no se infieren: §4.2 exige determinismo, y una heurística
# que decida "esto parece un identificador" falla en silencio con un export
# nuevo. Una cabecera que no case con ningún esquema se indexa entera, como
# siempre.
ESQUEMAS: tuple[EsquemaTabular, ...] = (
    EsquemaTabular(
        nombre="pubmed",
        firma=frozenset(
            {"PMID", "Title", "Authors", "Citation", "Journal/Book", "Publication Year"}
        ),
        indexables=("Title", "Authors", "Journal/Book", "Publication Year"),
    ),
    EsquemaTabular(
        nombre="litcovid",
        firma=frozenset({"pmid", "title", "journal"}),
        indexables=("title", "journal"),
    ),
)


def extraer(path: Path, fenomeno: int) -> Documento:
    """Extrae un ``Documento`` desde un CSV o un XLSX. Nunca lanza."""
    path = Path(path)
    formato = FORMATOS_SOPORTADOS.get(path.suffix.lower(), "csv")

    try:
        if formato == "xlsx":
            return _extraer_xlsx(path, fenomeno)
        return _extraer_csv(path, fenomeno)
    except Exception as exc:  # noqa: BLE001 - un archivo corrupto no tumba la corrida
        return documento_fallido(
            fuente=path.name,
            formato=formato,
            fenomeno=fenomeno,
            motivo=f"archivo tabular ilegible ({type(exc).__name__}): {exc}",
        )


# --- CSV ------------------------------------------------------------------------


def _extraer_csv(path: Path, fenomeno: int) -> Documento:
    texto, codificacion = _leer_texto(path)
    if texto is None:
        return documento_fallido(
            fuente=path.name,
            formato="csv",
            fenomeno=fenomeno,
            motivo=f"no se pudo decodificar el CSV con ninguna de {CODIFICACIONES}",
        )

    delimitador = _detectar_delimitador(texto)
    filas = list(csv.reader(io.StringIO(texto), delimiter=delimitador))
    cabecera = [normalizar_texto(celda) for celda in filas[0]] if filas else []

    jerarquia = Jerarquia()
    esquema = esquema_de(cabecera)
    registros, truncadas = _textos_de_filas(filas[1:], cabecera, esquema=esquema)
    bloques: list[Bloque | None] = [
        jerarquia.fila(texto, datos=datos) for texto, datos in registros
    ]

    meta: dict[str, Any] = {
        "columnas": cabecera,
        "delimitador": delimitador,
        "codificacion": codificacion,
        "n_filas": len(registros),
    }
    if esquema is not None:
        meta["esquema"] = esquema.nombre
        meta["columnas_indexadas"] = list(esquema.indexables)
    return _documento(path, "csv", fenomeno, bloques, meta, truncadas)


def _leer_texto(path: Path) -> tuple[str | None, str | None]:
    """Decodifica probando las codificaciones en orden fijo.

    Devuelve también cuál funcionó: un cambio de codificación cambia el texto, y
    sin dejarlo escrito no hay forma de saber después por qué dos corpus
    difieren.
    """
    crudo = path.read_bytes()
    if b"\x00" in crudo[:4096]:
        return None, None

    for codificacion in CODIFICACIONES:
        try:
            texto = crudo.decode(codificacion)
        except (UnicodeDecodeError, LookupError):
            continue
        # utf-8-sig acierta con y sin BOM; se informa del nombre corriente.
        return texto, "utf-8" if codificacion == "utf-8-sig" else codificacion
    return None, None


def _detectar_delimitador(texto: str) -> str:
    """Delimitador del CSV, con la coma como respaldo.

    El ``Sniffer`` falla con archivos de una sola columna o con comillas raras;
    ante la duda manda la coma, que es lo que produce el error más visible
    (una sola columna) en vez de uno silencioso.
    """
    muestra = texto[:MUESTRA_SNIFFER]
    try:
        return csv.Sniffer().sniff(muestra, delimiters=DELIMITADORES).delimiter
    except csv.Error:
        return ","


# --- XLSX -----------------------------------------------------------------------


def _extraer_xlsx(path: Path, fenomeno: int) -> Documento:
    # read_only para no cargar la hoja entera en memoria; data_only para que las
    # celdas con fórmula devuelvan su valor y no la fórmula.
    libro = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        jerarquia = Jerarquia()
        bloques: list[Bloque | None] = []
        columnas: dict[str, list[str]] = {}
        esquemas: dict[str, str] = {}
        truncadas = 0
        restantes = MAXIMO_FILAS

        for nombre in libro.sheetnames:
            filas = [
                [_celda(valor) for valor in fila]
                for fila in libro[nombre].iter_rows(values_only=True)
            ]
            if not filas:
                continue

            cabecera = [normalizar_texto(celda) for celda in filas[0]]
            esquema = esquema_de(cabecera)
            registros, sobrantes = _textos_de_filas(
                filas[1:], cabecera, tope=restantes, esquema=esquema
            )
            truncadas += sobrantes
            if not registros:
                continue

            # El título de la hoja se emite solo si tiene filas debajo, y antes
            # que ellas: una hoja vacía dejaría un bloque que solo dice "Datos",
            # y emitirlo después dejaría a las filas fuera de su propia sección.
            columnas[nombre] = cabecera
            if esquema is not None:
                esquemas[nombre] = esquema.nombre
            bloques.append(jerarquia.titulo(nombre, 1))
            bloques.extend(jerarquia.fila(texto, datos=datos) for texto, datos in registros)
            restantes -= len(registros)
    finally:
        libro.close()

    meta: dict[str, Any] = {
        "columnas": columnas,
        "hojas": list(columnas),
        "n_filas": sum(1 for bloque in bloques if bloque is not None and bloque.tipo == "fila"),
    }
    if esquemas:
        meta["esquema"] = esquemas
    return _documento(path, "xlsx", fenomeno, bloques, meta, truncadas)


def _celda(valor: Any) -> str:
    """Texto de una celda, con las fechas en ISO y sin decimales de adorno.

    Excel guarda las fechas como enteros seriales; ``openpyxl`` ya las convierte
    a ``datetime``, pero su ``str()`` arrastra la hora aunque no la haya.
    """
    if valor is None:
        return ""
    if isinstance(valor, datetime.datetime):
        return valor.date().isoformat() if valor.time() == datetime.time() else valor.isoformat(" ")
    if isinstance(valor, datetime.date):
        return valor.isoformat()
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor)


# --- común ----------------------------------------------------------------------


def esquema_de(cabecera: list[str]) -> EsquemaTabular | None:
    """El esquema declarado que reconoce esta cabecera, o ``None``.

    Se compara contra las columnas con nombre: la columna de índice sin
    cabecera que exporta pandas está en cuatro de los cinco CSV de PubMed y
    ausente en el quinto, y no debería partir el esquema en dos.
    """
    presentes = {normalizar_texto(str(columna)) for columna in cabecera}
    for esquema in ESQUEMAS:
        if esquema.firma <= presentes:
            return esquema
    return None


def _textos_de_filas(
    filas: list[list[str]],
    cabecera: list[str],
    tope: int = MAXIMO_FILAS,
    esquema: EsquemaTabular | None = None,
) -> tuple[list[tuple[str, dict[str, str]]], int]:
    """Serializa las filas con contenido. Devuelve también cuántas se dejaron fuera.

    Cada fila sale como ``(texto, descartados)``: lo que se indexa y lo que el
    esquema apartó. Serializar antes de construir bloques no es un rodeo: en
    XLSX hay que saber si una hoja tiene filas **antes** de emitir su título, o
    el título se emite igualmente y las filas acaban con un breadcrumb que no
    coincide con la pila de encabezados.
    """
    registros: list[tuple[str, dict[str, str]]] = []
    truncadas = 0

    for fila in filas:
        indexables, descartados = _repartir(list(_pares(cabecera, fila)), esquema)
        texto = serializar_registro(indexables)
        if not texto:
            continue
        if len(registros) >= tope:
            truncadas += 1
            continue
        registros.append((texto, descartados))

    return registros, truncadas


def _repartir(
    pares: list[tuple[str, str]], esquema: EsquemaTabular | None
) -> tuple[list[tuple[str, str]], dict[str, str]]:
    """Separa los pares en los que van al vector y los que van a metadata.

    Sin esquema declarado no se descarta nada: una cabecera que nadie reconoce
    se indexa entera, que es el comportamiento que había antes de que
    existieran los esquemas.
    """
    if esquema is None:
        return pares, {}

    valores = dict(pares)
    indexables = [
        (columna, valores[columna]) for columna in esquema.indexables if columna in valores
    ]
    descartados = {
        columna: valor
        for columna, valor in pares
        if columna not in esquema.indexables and normalizar_texto(str(valor))
    }
    return indexables, descartados


def _pares(cabecera: list[str], fila: list[str]):
    """Empareja cada celda con su columna, tolerando filas más largas o cortas.

    Una fila con más celdas que columnas no se descarta: sus celdas de más se
    numeran. Perder datos por un CSV mal formado es peor que indexar
    ``columna_7``.
    """
    for indice, celda in enumerate(fila):
        columna = cabecera[indice] if indice < len(cabecera) else f"columna_{indice + 1}"
        yield columna or f"columna_{indice + 1}", celda


def _documento(
    path: Path,
    formato: str,
    fenomeno: int,
    bloques: list[Bloque | None],
    meta: dict[str, Any],
    truncadas: int,
) -> Documento:
    errores = []
    if truncadas:
        meta["filas_truncadas"] = truncadas
        errores.append(
            f"dataset truncado en {MAXIMO_FILAS} filas: {truncadas} quedaron fuera del índice"
        )
    if not any(bloques):
        errores.append("el archivo no tiene filas con contenido")

    return construir_documento(
        fuente=path.name,
        formato=formato,
        fenomeno=fenomeno,
        bloques=bloques,
        meta=meta,
        errores=errores,
    )
